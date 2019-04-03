from openpyxl import load_workbook, Workbook
from datetime import timezone,datetime
import filebook
import database
import ecxel

#Создаем копию шаблонов адресов, определяем её путь 
url_adress=filebook.sadres()

#Загрузка книги данных
mainfile = load_workbook('main.xlsx')
wbadress = load_workbook(url_adress)

#Инициализация листов
sheetdata = mainfile["data"]
sheetinfo = mainfile["info"]
wsa = wbadress["adress_template"]
wsa_copy = wsa

#Заполняем список данными
data = tuple(sheetdata.values)
db = database.createdb(data)

#Импорт данных, 
all_adress = sheetinfo ['K2:K200']
for ls in all_adress:
    if (ls[0].value != None):
        #Создает копии листа
        ws_time = wbadress.copy_worksheet(wsa_copy)
        title = str(ls[0].value)
        ws_time.title = title
        ws = wbadress[title]
        workspace  = ws['I7:H34']

        sql = """SELECT date, name, name_object, quantity, 
        price, sum, discount, end_sum, comment
        from test
        WHERE adres = ?
        """ 
        db.execute(sql, [(str(ls[0].value))] )
        select = tuple(db.fetchall())
        print(str(ls[0].value) + " данные по адресам загруженны")

        #Установка Значений
        ws_time['A6'] = str(ls[0].value)
        try:
            for i in select:
                workspace[0][1].value = i[1]  
        except IndexError:
            print("Нет Данных")
            




        


 








sql = """SELECT date, name, name_object, quantity, 
price, sum, discount, end_sum, comment
from test
WHERE adres = 'Мост5(+)'
"""
db.execute(sql)
#for str in db.fetchall():
#    print (str)

wbadress.save(url_adress)
db.execute(" DROP TABLE test " )


